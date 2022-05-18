import os
from packages.dirs import Dirs
from packages.config import Config
from openpyxl import load_workbook
from app.classes.Logging import Logging

class Input():
    files = None
    current_file = ''
    current_rayon = ''

    def __init__(self):
        all_files = self.showFiles()
        self.files = self.filterFiles(all_files)
        self.input_dir = Dirs().get('input')
        self.rayons = Config('rayons').get()
        self.from_row = Config().get('input_from_row')

    # Получаем список файлов в папке input
    def showFiles(self):
        input_dir = Dirs().get('input')
        return [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f))]

    # Удаляем лишние файлы из списка файлов input
    def filterFiles(self, files):
        filter_files = []
        for file in files:
            if (file.endswith('.xlsx')):
                filter_files.append(file)
        return filter_files

    # Получаем ID района по заголовку
    def getRayonFromTitle(self, title):
        rows = None
        for rayon in self.rayons:
            if rayon['search_not'] == "":
                if (rayon['search'] in title.lower()):
                    rows = dict({
                        "id":rayon['id'],
                        "row_fed":rayon['row_fed'],
                        "row_vet":rayon['row_vet'],
                        "row_reab":rayon['row_reab'],
                        "row_mnogodet":rayon['row_mnogodet'],
                        "row_35":rayon['row_35'],
                        "row_spec":rayon['row_spec']
                    })
            else:
                if (rayon['search'] in title.lower()) and (rayon['search_not'] not in title.lower()):
                    rows = dict({
                        "id":rayon['id'],
                        "row_fed":rayon['row_fed'],
                        "row_vet":rayon['row_vet'],
                        "row_reab": rayon['row_reab'],
                        "row_mnogodet":rayon['row_mnogodet'],
                        "row_35":rayon['row_35'],
                        "row_spec":rayon['row_spec']
                    })

        if rows == None:
            error_text = 'Ошибка!!! Не удалось определить район по следующему заголовку: "{}"'.format(title)
            Logging().error(error_text)
            exit(error_text)
        return rows

    # Получаем информацию о типе выплаты содержащую С и ПО какую строку брать
    def getVplRange(self, ws):
        vpl = []

        max_row = ws.max_row
        merged_all = ws.merged_cells

        size_block = []
        temp_vpl = dict()

        for row in range(self.from_row, max_row):

            # Находим с какой строки начинается новый блок
            cell_a = "{}{}".format('A', row)
            if cell_a in merged_all:
                temp_vpl['from'] = row

            # Находим на какой строке заканчивается блок
            cell_c = "{}{}".format('C', row)
            try:
                if 'Итого по выплате' in ws[cell_c].value:
                    temp_vpl['to'] = row
            except:
                pass

            # Сохраняем весь блок
            if len(temp_vpl) > 1:
                size_block.append(temp_vpl)
                temp_vpl = dict()

        return size_block

    # Получаем информацию об одной выплате
    def getOneVpl(self, ws, range_one):
        dataBlock = dict()

        cell_title = "{}{}".format('A', range_one['from'])
        title = ws[cell_title].value

        dataBlock['title'] = title

        cell_total_col = "{}{}".format('F', range_one['to'])
        dataBlock['total_col'] = ws[cell_total_col].value

        cell_total_sum_row = "{}{}".format('G', range_one['to'])
        cell_total_sum = ws[cell_total_sum_row].value

        if not (cell_total_sum == None):
            cell_total_sum = float(cell_total_sum.replace(',', '.'))
            cell_total_sum = float("{0:.2f}".format(cell_total_sum))

        dataBlock['total_sum'] = cell_total_sum

        dataBlock['sb_col'] = 0
        dataBlock['sb_sum'] = 0
        for row in range(range_one['from'], range_one['to'] + 1):
            cell_vid_row = "{}{}".format('L', row)
            cell_vid = ws[cell_vid_row].value
            if cell_vid == 'сб/б':
                cell_sb_raw = "{}{}".format('G', row)
                cell_sb_str = ws[cell_sb_raw].value.strip()
                cell_sb = cell_sb_str.replace(',', '.')

                cell_col_raw = "{}{}".format('F', row)
                cell_col = ws[cell_col_raw].value.strip()

                try:
                    dataBlock['sb_col'] = dataBlock['sb_col'] + int(cell_col)
                    dataBlock['sb_sum'] = dataBlock['sb_sum'] + float(cell_sb)
                except:
                    pass

            dataBlock['sb_sum'] = float("{0:.2f}".format(dataBlock['sb_sum']))

        return dataBlock

    # Добавляем ID к одной выплате
    def getVplId(self, vpl):
        title = vpl['title'].strip()
        vpls = Config('vpl').get()

        vpl_meta = None
        for vpl in vpls:
            for like in vpl['like']:
                if like.lower() in title.lower():
                    vpl_meta = dict({
                        "id": vpl['id'],
                        "type": vpl['type'],
                        "page_title": vpl['page_title'],
                    })

        if vpl_meta == None:
            error_text = 'Ошибка!!! Не удалось определить ID выплаты по заголовку: "{}"'.format(title)
            Logging().error(error_text)

        return vpl_meta


    # Добавляем ID к одной выплате
    def getTemplate(self, vpl_id):
        templates = Config('template').get()

        vpl_template = None
        for template in templates:
            if template['vpl_id'] == vpl_id:
                    vpl_template = dict({
                        "sb_col": template['sb_col'],
                        "sb_sum": template['sb_sum'],
                        "total_col": template['total_col'],
                        "total_sum": template['total_sum'],
                    })
        if vpl_template == None:
            pass
            # error_text = 'Ошибка!!! Не удалось определить ID выплаты по заголовку: "{}"'.format(vpl_id)
            # Logging().error(error_text)

        return vpl_template

    def filterVplOnMounth(self, vpl):
        title = vpl['title'].strip()

        app_config = Config().get()
        user_config = Config('user_config').get()
        user_mounth = user_config['mounth']
        user_year = user_config['year']

        search_string = ''
        if not ((user_mounth == 0) or (user_mounth == "")):
            for mounth in app_config['mounth']:
                if mounth['id'] == user_mounth:
                    search_string = str(mounth['title'])

        if not (user_year == 0) or (user_year == ""):
            if not (search_string == ''):
                search_string = search_string + ' ' + str(user_year)
            else:
                search_string = str(user_year)

        if not (search_string == None):
            if search_string.lower() in title.lower():
                return True
            return False

    def filterVpl(self, vpl):
        title = vpl['title'].strip()
        del_vpl_list = Config('delete_vpls').get()
        for d_vpl in del_vpl_list:
            if d_vpl.lower() in title.lower():
                return False
        return True

    # Получаем информацию о всех выплатах
    def getVpls(self, ws):
        vpl = []
        ranges = self.getVplRange(ws)
        for range_one in ranges:
            one_vpl = self.getOneVpl(ws, range_one)
            f_vpl = self.filterVpl(one_vpl)
            f_vpl_on_mounth = self.filterVplOnMounth(one_vpl)


            if (f_vpl == True) and (f_vpl_on_mounth == True):
                vpl_meta = self.getVplId(one_vpl)
                if not (vpl_meta == None):
                    one_vpl['meta'] = vpl_meta
                    one_vpl['template'] = self.getTemplate(vpl_meta['id'])

                    vpl.append(one_vpl)
            else:
                error_text = 'Внимание!!! Следующая выплата была отфильтрована: "{}". Учреждение: "{}". Файл: "{}"'.format(one_vpl, self.current_rayon, self.current_file)
                Logging().warning(error_text)
        return vpl


    # Получаем содержимое одного файла
    def getOne(self, file):
        data = dict()

        file_path = os.path.join(self.input_dir, file)
        wb = load_workbook(filename=file_path)
        ws = wb.worksheets[0]

        # Получаем название Района
        title_cell_raw = "{}{}".format('A', 2)
        title = ws[title_cell_raw].value.strip()
        data['file_name'] = file
        data['title'] = title

        self.current_file = file
        self.current_rayon = title

        # Получаем номера строк по названию
        data['meta'] = self.getRayonFromTitle(title)

        # Добавляем все выплаты
        data['vpls'] = self.getVpls(ws)

        wb.close()
        return data

    # Получаем список всех файлов
    def getAll(self):
        data = []
        for file in self.files:
            file_content = self.getOne(file)
            data.append(file_content)
        return data