import os
from packages.dirs import Dirs
from packages.config import Config
from openpyxl import load_workbook
from app.classes.Logging import Logging

class Input():
    files = None

    def __init__(self):
        all_files = self.showFiles()
        self.files = self.filterFiles(all_files)
        self.input_dir = Dirs().get('input')
        self.rayons = Config('rayons').get()
        self.from_row = Config().get('input_from_row')

    # Получаем список файлов в папке input
    def showFiles(self):
        input_dir = Dirs().get('input')
        return [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f))]

    # Удаляем лишние файлы из списка файлов input
    def filterFiles(self, files):
        filter_files = []
        for file in files:
            if (file.endswith('.xlsx')):
                filter_files.append(file)
        return filter_files

    # Получаем ID района по заголовку
    def getRayonFromTitle(self, title):
        rayon_id = 0
        for rayon in self.rayons:
            if rayon['search_not'] == "":
                if (rayon['search'] in title.lower()):
                    rayon_id = rayon['id']
            else:
                if (rayon['search'] in title.lower()) and (rayon['search_not'] not in title.lower()):
                    rayon_id = rayon['id']

        if rayon_id == 0:
            error_text = 'Ошибка!!! Не удалось определить район по следующему заголовку: "{}"'.format(title)
            Logging().error(error_text)
            exit(error_text)
        return rayon_id

    # Получаем информацию о типе выплаты содержащую С и ПО какую строку брать
    def getVplRange(self, ws):
        vpl = []

        max_row = ws.max_row
        merged_all = ws.merged_cells

        size_block = []
        temp_vpl = dict()

        for row in range(self.from_row, max_row):

            # Находим с какой строки начинается новый блок
            cell_a = "{}{}".format('A', row)
            if cell_a in merged_all:
                temp_vpl['from'] = row

            # Находим на какой строке заканчивается блок
            cell_c = "{}{}".format('C', row)
            try:
                if 'Итого по выплате' in ws[cell_c].value:
                    temp_vpl['to'] = row
            except:
                pass

            # Сохраняем весь блок
            if len(temp_vpl) > 1:
                size_block.append(temp_vpl)
                temp_vpl = dict()

        return size_block

    # Получаем информацию об одной выплате
    def getOneVpl(self, ws, range_one):
        dataBlock = dict()

        cell_title = "{}{}".format('A', range_one['from'])
        title = ws[cell_title].value

        dataBlock['title'] = title

        cell_total_col = "{}{}".format('F', range_one['to'])
        dataBlock['total_col'] = ws[cell_total_col].value

        cell_total_sum = "{}{}".format('G', range_one['to'])
        dataBlock['total_sum'] = ws[cell_total_sum].value

        dataBlock['sb_col'] = 0
        dataBlock['sb_sum'] = 0
        for row in range(range_one['from'], range_one['to'] + 1):
            cell_vid_row = "{}{}".format('L', row)
            cell_vid = ws[cell_vid_row].value
            if cell_vid == 'сб/б':
                cell_sb_raw = "{}{}".format('G', row)
                cell_sb_str = ws[cell_sb_raw].value.strip()
                cell_sb = cell_sb_str.replace(',', '.')

                cell_col_raw = "{}{}".format('F', row)
                cell_col = ws[cell_col_raw].value.strip()

                try:
                    dataBlock['sb_col'] = dataBlock['sb_col'] + int(cell_col)
                    dataBlock['sb_sum'] = dataBlock['sb_sum'] + float(cell_sb)
                except:
                    pass

            dataBlock['sb_sum'] = float("{0:.2f}".format(dataBlock['sb_sum']))

        return dataBlock

    # Добавляем ID к одной выплате
    def getVplId(self, vpl):
        title = vpl['title']
        vpls = Config('vpl').get()

        vpl_id = 0
        for vpl in vpls:
            for like in vpl['like']:
                if like.lower() in title.lower():
                    vpl_id = vpl['id']

        if vpl_id == 0:
            error_text = 'Ошибка!!! Не удалось определить ID выплаты по заголовку: "{}"'.format(title)
            Logging().error(error_text)
            exit(error_text)

        return vpl_id

    # Получаем информацию о всех выплатах
    def getVpls(self, ws):
        vpl = []
        ranges = self.getVplRange(ws)
        for range_one in ranges:
            one_vpl = self.getOneVpl(ws, range_one)
            vpl_id = self.getVplId(one_vpl)
            one_vpl['id'] = vpl_id
            vpl.append(one_vpl)
        return vpl


    # Получаем содержимое одного файла
    def getOne(self, file):
        data = dict()

        file_path = os.path.join(self.input_dir, file)
        wb = load_workbook(filename=file_path)
        ws = wb.worksheets[0]

        # Получаем название Района
        title_cell_raw = "{}{}".format('A', 2)
        title = ws[title_cell_raw].value.strip()

        # Получаем id района по названию
        rayon_id = self.getRayonFromTitle(title)

        data['id'] = rayon_id
        data['title'] = title
        # Добавляем все выплаты
        data['vpls'] = self.getVpls(ws)

        print(data)
        exit()

        wb.close()
        return data

    # Получаем список всех файлов
    def getAll(self):
        data = []
        for file in self.files:
            file_content = self.getOne(file)
            data.append(file_content)
        return data