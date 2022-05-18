import os
from packages.dirs import Dirs
from packages.config import Config
from openpyxl import load_workbook, Workbook

class Input():
    ext = 'xlsx'
    files = None

    def __init__(self):
        all_files = self.showFiles()
        self.files = self.filterFiles(all_files)
        self.input_dir = Dirs().get('input')
        self.rayons = Config('rayons').get()
        self.from_row = Config().get('input_from_row')

    def showFiles(self):
        input_dir = Dirs().get('input')
        return [f for f in os.listdir(input_dir) if os.path.isfile(os.path.join(input_dir, f))]

    def filterFiles(self, files):
        filter_files = []
        for file in files:
            if (file.endswith("." + self.ext)):
                filter_files.append(file)
        return filter_files

    def getIndexFromTitle(self, title):
        for rayon in self.rayons:
            if rayon['search_not'] == "":
                if (rayon['search'] in title.lower()):
                    return {"rayon_id":rayon['id'], "fed_row":rayon['fed_row']}
            else:
                if (rayon['search'] in title.lower()) and (rayon['search_not'] not in title.lower()):
                    return {"rayon_id":rayon['id'], "fed_row":rayon['fed_row']}
        return {"rayon_id":0, "fed_row":0}

    def getBlocks(self, ws):
        vpl = []

        max_row = ws.max_row
        merged_all = ws.merged_cells

        size_block = []
        temp_vpl = dict()

        for row in range(self.from_row, max_row):

            # Находим с какой строки начинается новый блок
            cell_a = "{}{}".format('A', row)
            if cell_a in merged_all:
                temp_vpl['from'] = row

            # Находим на какой строке заканчивается блок
            cell_c = "{}{}".format('C', row)
            try:
                if 'Итого по выплате' in ws[cell_c].value:
                    temp_vpl['to'] = row
            except:
                pass

            # Сохраняем весь блок
            if len(temp_vpl) > 1:
                size_block.append(temp_vpl)
                temp_vpl = dict()

        return size_block

    # Обязательно сделать оповещение, если не сходится справочник
    def setVplIds(self, title):
        vpls = Config('vpl').get()
        for vpl in vpls:
            for like in vpl['like']:
                if like.lower() in title.lower():
                    return vpl['id']
        return 0

    # Получаем данные всех выплат
    def getVpl(self, ws):
        vpl = []
        blocks = self.getBlocks(ws)
        for block in blocks:
            dataBlock = dict()

            cell_title = "{}{}".format('A', block['from'])
            title = ws[cell_title].value

            vpl_id = self.setVplIds(title)
            if vpl_id == 0:
                print(0, title)

            dataBlock['id'] = vpl_id
            dataBlock['title'] = title

            cell_total_col = "{}{}".format('F', block['to'])
            dataBlock['total_col'] = ws[cell_total_col].value

            cell_total_sum = "{}{}".format('G', block['to'])
            dataBlock['total_sum'] = ws[cell_total_sum].value

            dataBlock['sb_col'] = 0
            dataBlock['sb_sum'] = 0
            for row in range(block['from'], block['to'] + 1):
                cell_vid_row = "{}{}".format('L', row)
                cell_vid = ws[cell_vid_row].value
                if cell_vid == 'сб/б':
                    cell_sb_raw = "{}{}".format('G', row)
                    cell_sb_str = ws[cell_sb_raw].value.strip()
                    cell_sb = cell_sb_str.replace(',', '.')

                    cell_col_raw = "{}{}".format('F', row)
                    cell_col = ws[cell_col_raw].value.strip()

                    try:
                        dataBlock['sb_col'] = dataBlock['sb_col'] + int(cell_col)
                        dataBlock['sb_sum'] = dataBlock['sb_sum'] + float(cell_sb)
                    except:
                        pass

                dataBlock['sb_sum'] = float("{0:.2f}".format(dataBlock['sb_sum']))
            vpl.append(dataBlock)
        return vpl


    # --------------------------

    def getOne(self, file):
        data = dict()

        file_path = os.path.join(self.input_dir, file)
        wb = load_workbook(filename=file_path)
        ws = wb.worksheets[0]

        # Получаем название Района
        title_cell_raw = "{}{}".format('A', 2)
        title = ws[title_cell_raw].value.strip()

        rayons = self.getIndexFromTitle(title)

        vpl = self.getVpl(ws)

        data['rayon_id'] = rayons['rayon_id']
        data['fed_row'] = rayons['fed_row']
        data['title'] = title
        data['vpl'] = vpl

        return data

    def getAll(self):
        data = []
        for file in self.files:
            file_content = self.getOne(file)
            data.append(file_content)
        return data