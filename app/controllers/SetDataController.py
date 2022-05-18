import os
from shutil import copyfile
from app.models.Output import Output
from packages.config import Config
from packages.dirs import Dirs
from openpyxl import load_workbook, Workbook
from app.classes.Logging import Logging

class SetDataController():

    output_file = None
    wb = None

    def __init__(self):
        template_dir = Dirs().get('template')
        output_dir = Dirs().get('output')

        file_name = Config().get('output_file')
        template_file = os.path.join(template_dir, file_name)
        output_file = os.path.join(output_dir, file_name)
        copyfile(template_file, output_file)

        # Загружаем объект excel в свойства
        self.wb = load_workbook(filename=output_file)
        self.output_file = output_file

    def setContent(self, all):
        for one in all:
            print(one['title'], ' - ', one['file_name'])
            self.setOneContent(one)

    def getTypeName(self, id):
        type_name = None
        vpl_types = Config('vpl_types').get()
        for type in vpl_types:
            if type['id'] == id:
                type_name = type['type']
        if type_name == None:
            error_text = 'Ошибка!!! Не удалось определить название типа выплаты по ID: "{}"'.format(id)
            Logging().error(error_text)
            exit(error_text)
        return type_name

    def getNamePage(self, id):
        page_name = None
        vpl_types = Config('vpl_types').get()
        for type in vpl_types:
            if type['id'] == id:
                page_name = type['page_title']
        if page_name == None:
            error_text = 'Ошибка!!! Не удалось определить название типа выплаты по ID: "{}"'.format(id)
            Logging().error(error_text)
            exit(error_text)
        return page_name

    def getTemplate(self, vpl_id):
        tmp = None
        template = Config('template').get()

        for template_one in template:
            if template_one['vpl_id'] == vpl_id:
                tmp = template_one

        if tmp == None:
            error_text = 'Ошибка!!! Не удалось определить шаблон выплаты по ID выплаты: "{}"'.format(vpl_id)
            Logging().error(error_text)
            exit(error_text)
        return tmp

    def setOneData(self, vpl, meta_vpl, row):
        page_title = meta_vpl['page_title']
        ws = self.wb[page_title]

        if 'template' in vpl:
            template_vpl = vpl['template']

            sb_col_raw = "{}{}".format(template_vpl['sb_col'], row)
            sb_sum_raw = "{}{}".format(template_vpl['sb_sum'], row)
            total_col_raw = "{}{}".format(template_vpl['total_col'], row)
            total_sum_raw = "{}{}".format(template_vpl['total_sum'], row)

            old_sb_col = ws[sb_col_raw].value
            old_sb_sum = ws[sb_sum_raw].value
            old_total_col = ws[total_col_raw].value
            old_total_sum = ws[total_sum_raw].value

            print(old_sb_col, vpl)

            if (not (old_sb_col == None) and (old_sb_col > 0)):
                if type(old_sb_col) == str:
                    old_sb_col = old_sb_col.strip()
                    old_sb_col = int(old_sb_col)
                ws[sb_col_raw] = old_sb_col + int(vpl['sb_col'])
            else:
                ws[sb_col_raw] = int(vpl['sb_col'])

            if (not (old_sb_sum == None) and (old_sb_sum > 0)):
                if type(old_sb_sum) == str:
                    old_sb_sum = old_sb_sum.strip()
                    old_sb_sum = float(old_sb_sum.replace(',', '.'))

                old_sb_sum = float("{0:.2f}".format(old_sb_sum))
                ws[sb_sum_raw] = old_sb_sum + float(vpl['sb_sum'])
            else:
                ws[sb_sum_raw] = float(vpl['sb_sum'])

            if (not (old_total_col == None) and (old_total_col > 0)):
                if type(old_total_col) == str:
                    old_total_col = old_total_col.strip()
                    old_total_col = int(old_total_col)
                ws[total_col_raw] = old_total_col + int(vpl['total_col'])
            else:
                ws[total_col_raw] = int(vpl['total_col'])

            if (not (old_total_sum == None) and (old_total_sum > 0)):
                if type(old_total_sum) == str:
                    old_total_sum = old_total_sum.strip()
                    old_total_sum = float(old_total_sum.replace(',', '.'))

                old_total_sum = float("{0:.2f}".format(old_total_sum))
                ws[total_sum_raw] = old_total_sum + float(vpl['total_sum'])
            else:
                ws[total_sum_raw] = float(vpl['total_sum'])


    def setData(self, vpl, data):
        row = 0

        if 'meta' in vpl:
            meta_vpl = vpl['meta']
            meta_data = data['meta']

            if meta_vpl['type'] == 'fed':
                row = meta_data['row_fed']
            elif meta_vpl['type'] == 'reab':
                row = meta_data['row_reab']
            elif meta_vpl['type'] == 'spec':
                row = meta_data['row_spec']
            elif meta_vpl['type'] == 'mnog':
                row = meta_data['row_mnogodet']
            elif meta_vpl['type'] == '35_6':
                row = meta_data['row_35']
            elif meta_vpl['type'] == 'vet':
                row = meta_data['row_vet']
            else:
                pass
                #error_text = 'Не удалось определить тип выплаты'
                #Logging().error(error_text)

            if not (row == 0):
               self.setOneData(vpl, meta_vpl, row)

        self.wb.save(self.output_file)


    def getVplType(self, vpl):
        vpl_type = 0

        vpl_list = Config('vpl').get()
        for item in vpl_list:
            if vpl['id'] == item['id']:
                vpl_type = item['type_id']

        if vpl_type == 0:
            error_text = 'Ошибка!!! Не удалось определить тип выплаты по заголовку: "{}"'.format(vpl['title'])
            Logging().error(error_text)
            exit(error_text)

        return vpl_type


    def setOneContent(self, data):
        for vpl in data['vpls']:
            self.setData(vpl, data)

    def mergeMeta(self, vpl):
        template = Config('template').get()
        for template_vpl in template:
            if template_vpl['vpl_id'] == vpl['id']:
                vpl['meta'] = template_vpl
        return vpl